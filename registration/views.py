from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from .models import Registration, CourseClass, StudentProgress, UserProfile, AppSetting
from .forms import RegistrationForm
import csv
from datetime import datetime
import openpyxl

def landing_view(request):
    reg_setting, _ = AppSetting.objects.get_or_create(key='registration_locked', defaults={'value_bool': False})
    is_locked = reg_setting.value_bool
    return render(request, 'registration/landing.html', {'is_locked': is_locked})

def register_view(request):
    is_locked = AppSetting.objects.filter(key='registration_locked', value_bool=True).exists()
    if is_locked:
        return redirect('landing')
        
    if request.method == 'POST':
        form = RegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            registration = form.save(commit=False)
            registration.is_paid = False # Force to False so admin has to verify screenshot
            registration.save()
            return redirect('register_success', pk=registration.pk)
    else:
        form = RegistrationForm()
    return render(request, 'registration/register.html', {'form': form})

def register_success(request, pk):
    reg = get_object_or_404(Registration, pk=pk)
    return render(request, 'registration/success.html', {'reg': reg})

@login_required
def dashboard_router(request):
    try:
        profile = request.user.profile
        if profile.role == 'ADMIN':
            return redirect('admin_dashboard')
        elif profile.role == 'MENTOR':
            return redirect('mentor_dashboard')
        else:
            return redirect('student_dashboard')
    except:
        if request.user.is_superuser:
            return redirect('admin_dashboard')
        return redirect('landing')

@login_required
def admin_dashboard_view(request):
    if not request.user.is_superuser and getattr(request.user, 'profile', None) and request.user.profile.role != 'ADMIN':
        return redirect('landing')
        
    registrations = Registration.objects.all().order_by('-created_at')
    total_registered = registrations.count()
    total_paid = registrations.filter(is_paid=True).count()
    
    mentors = UserProfile.objects.filter(role='MENTOR')
    course_classes = CourseClass.objects.all().order_by('order')
    students = UserProfile.objects.filter(role='STUDENT')
    
    reg_setting, _ = AppSetting.objects.get_or_create(key='registration_locked', defaults={'value_bool': False})
    is_locked = reg_setting.value_bool
    
    context = {
        'registrations': registrations,
        'total_registered': total_registered,
        'total_paid': total_paid,
        'mentors': mentors,
        'course_classes': course_classes,
        'students': students,
        'is_locked': is_locked,
    }
    return render(request, 'registration/dashboard.html', context)

@login_required
def admin_toggle_lock_view(request):
    if not request.user.is_superuser and getattr(request.user, 'profile', None) and request.user.profile.role != 'ADMIN':
        return redirect('landing')
        
    if request.method == 'POST':
        is_locked_val = request.POST.get('is_locked') == 'on'
        AppSetting.objects.update_or_create(key='registration_locked', defaults={'value_bool': is_locked_val})
        
    return redirect('admin_dashboard')


@login_required
def admin_manage_class_view(request):
    if not request.user.is_superuser and getattr(request.user, 'profile', None) and request.user.profile.role != 'ADMIN':
        return redirect('landing')
        
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            raw_vid = request.POST.get('youtube_video_id', '').strip()
            
            # Safely extract YouTube ID if user pasted full URL
            import re
            vid_id = raw_vid
            
            # Match standard youtube.com/watch?v=ID or youtu.be/ID
            match = re.search(r'(?:v=|youtu\.be/|embed/)([^&?/\s]{11})', raw_vid)
            if match:
                vid_id = match.group(1)
            elif len(raw_vid) == 11:
                vid_id = raw_vid
            else:
                # Fallback, just try to take the last 11 characters if it's a weird url, or just save as is
                vid_id = raw_vid[-11:] if len(raw_vid) > 11 else raw_vid
                
            CourseClass.objects.create(
                title=request.POST.get('title'),
                youtube_video_id=vid_id,
                order=request.POST.get('order'),
                description=request.POST.get('description', '')
            )
        elif action == 'edit':
            class_id = request.POST.get('class_id')
            raw_vid = request.POST.get('youtube_video_id', '').strip()
            
            import re
            vid_id = raw_vid
            match = re.search(r'(?:v=|youtu\.be/|embed/)([^&?/\s]{11})', raw_vid)
            if match:
                vid_id = match.group(1)
            elif len(raw_vid) == 11:
                vid_id = raw_vid
            else:
                vid_id = raw_vid[-11:] if len(raw_vid) > 11 else raw_vid
                
            CourseClass.objects.filter(id=class_id).update(
                title=request.POST.get('title'),
                youtube_video_id=vid_id,
                order=request.POST.get('order'),
                description=request.POST.get('description', '')
            )
        elif action == 'delete':
            class_id = request.POST.get('class_id')
            CourseClass.objects.filter(id=class_id).delete()
            
    return redirect('admin_dashboard')

@login_required
def admin_manage_mentor_view(request):
    if not request.user.is_superuser and getattr(request.user, 'profile', None) and request.user.profile.role != 'ADMIN':
        return redirect('landing')
        
    if request.method == 'POST':
        name = request.POST.get('name')
        mobile = request.POST.get('mobile')
        password = request.POST.get('password')
        
        if not User.objects.filter(username=mobile).exists():
            user = User.objects.create_user(username=mobile, password=password, first_name=name)
            UserProfile.objects.create(user=user, role='MENTOR')
            
    return redirect('admin_dashboard')

@login_required
def admin_assign_students_view(request):
    if not request.user.is_superuser and getattr(request.user, 'profile', None) and request.user.profile.role != 'ADMIN':
        return redirect('landing')
        
    if request.method == 'POST':
        action = request.POST.get('action')
        student_ids = request.POST.getlist('student_ids') # These are now Registration IDs
        
        if not student_ids:
            return redirect('admin_dashboard')
            
        if action == 'assign':
            mentor_id = request.POST.get('mentor_id')
            if mentor_id:
                mentor_profile = get_object_or_404(UserProfile, id=mentor_id, role='MENTOR')
                UserProfile.objects.filter(registration__id__in=student_ids, role='STUDENT').update(mentor=mentor_profile)
        
        elif action == 'delete':
            # Bulk delete
            # Delete associated User accounts first (cascades to UserProfile)
            users_to_delete = User.objects.filter(profile__registration__id__in=student_ids)
            users_to_delete.delete()
            # Delete registrations
            for reg in Registration.objects.filter(id__in=student_ids):
                if reg.screenshot and os.path.isfile(reg.screenshot.path):
                    os.remove(reg.screenshot.path)
                reg.delete()
                
    return redirect('admin_dashboard')

import os

@login_required
def admin_verify_payment_view(request):
    if not request.user.is_superuser and getattr(request.user, 'profile', None) and request.user.profile.role != 'ADMIN':
        return redirect('landing')
        
    if request.method == 'POST':
        reg_id = request.POST.get('reg_id')
        reg = get_object_or_404(Registration, id=reg_id)
        if not reg.is_paid:
            reg.is_paid = True
            reg.save() # This triggers post_save to create the User account
            
            # Delete the screenshot to reduce storage
            if reg.screenshot:
                if os.path.isfile(reg.screenshot.path):
                    os.remove(reg.screenshot.path)
                reg.screenshot = None
                reg.save()
                
    return redirect('admin_dashboard')

@login_required
def admin_edit_registration_view(request):
    if not request.user.is_superuser and getattr(request.user, 'profile', None) and request.user.profile.role != 'ADMIN':
        return redirect('landing')
        
    if request.method == 'POST':
        reg_id = request.POST.get('reg_id')
        reg = get_object_or_404(Registration, id=reg_id)
        
        old_mobile = reg.mobile
        reg.name = request.POST.get('name', reg.name)
        reg.house_name = request.POST.get('house_name', reg.house_name)
        reg.place = request.POST.get('place', reg.place)
        reg.post = request.POST.get('post', reg.post)
        reg.district = request.POST.get('district', reg.district)
        reg.mobile = request.POST.get('mobile', reg.mobile)
        reg.whatsapp = request.POST.get('whatsapp', reg.whatsapp)
        
        reg.save()
        
        # If mobile changed and they are already a user, update their login username
        if old_mobile != reg.mobile and reg.is_paid:
            try:
                user = User.objects.get(username=old_mobile)
                user.username = reg.mobile
                user.save()
            except User.DoesNotExist:
                pass
                
    return redirect('admin_dashboard')

@login_required
def export_excel_view(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=registrations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Registrations'
    
    columns = [
        'ID', 'Name', 'House Name', 'Place', 'Post', 'District',
        'Mobile', 'WhatsApp', 'Paid', 'Transaction Time', 'Transaction ID', 'Created At'
    ]
    row_num = 1
    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        
    for reg in Registration.objects.all().order_by('-created_at'):
        row_num += 1
        row = [
            f"APP-{reg.application_number}", reg.name, reg.house_name, reg.place, reg.post, reg.district,
            reg.mobile, reg.whatsapp, 'Yes' if reg.is_paid else 'No',
            reg.transaction_time_and_date.strftime('%Y-%m-%d %H:%M:%S') if reg.transaction_time_and_date else '',
            reg.transaction_id,
            reg.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            
    workbook.save(response)
    return response

@login_required
def student_dashboard_view(request):
    if getattr(request.user, 'profile', None) and request.user.profile.role != 'STUDENT':
        return redirect('landing')
        
    classes = CourseClass.objects.all().order_by('order')
    progress_list = StudentProgress.objects.filter(student=request.user)
    
    completed_class_ids = set(p.course_class.id for p in progress_list if p.is_completed)
    
    class_data = []
    is_unlocked = True
    
    for c in classes:
        completed = c.id in completed_class_ids
        class_data.append({
            'class': c,
            'is_unlocked': is_unlocked,
            'is_completed': completed
        })
        is_unlocked = completed
        
    return render(request, 'registration/student_dashboard.html', {'class_data': class_data})

@login_required
def classroom_view(request, class_id):
    if getattr(request.user, 'profile', None) and request.user.profile.role != 'STUDENT':
        return redirect('landing')
        
    course_class = get_object_or_404(CourseClass, id=class_id)
    
    # Check if unlocked
    if course_class.order > 1:
        prev_class = CourseClass.objects.filter(order=course_class.order - 1).first()
        if prev_class:
            prev_progress = StudentProgress.objects.filter(student=request.user, course_class=prev_class, is_completed=True).exists()
            if not prev_progress:
                return HttpResponse("Please watch the previous class first.", status=403)
                
    is_already_completed = StudentProgress.objects.filter(student=request.user, course_class=course_class, is_completed=True).exists()
                
    if request.method == 'POST' and not is_already_completed:
        progress, created = StudentProgress.objects.get_or_create(student=request.user, course_class=course_class)
        progress.is_completed = True
        progress.save()
        return redirect('student_dashboard')
        
    return render(request, 'registration/classroom.html', {'course_class': course_class, 'is_already_completed': is_already_completed})

@login_required
def mentor_dashboard_view(request):
    if getattr(request.user, 'profile', None) and request.user.profile.role != 'MENTOR':
        return redirect('landing')
        
    total_classes = CourseClass.objects.count()
    student_data = []
    
    try:
        profile = request.user.profile
        assigned_students = profile.mentees.all()
        
        for student_profile in assigned_students:
            user = student_profile.user
            completed_count = StudentProgress.objects.filter(student=user, is_completed=True).count()
            last_progress = StudentProgress.objects.filter(student=user, is_completed=True).order_by('-completed_at').first()
            
            student_data.append({
                'name': user.first_name,
                'mobile': user.username,
                'completed_count': completed_count,
                'total_classes': total_classes,
                'last_activity': last_progress.completed_at if last_progress else None
            })
    except:
        pass
        
    return render(request, 'registration/mentor_dashboard.html', {'student_data': student_data})
